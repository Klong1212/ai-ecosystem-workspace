"""
OpenAPI to CSV/Excel Converter

แปลง openapi.json จาก FastAPI ให้เป็น CSV หรือ Excel สำหรับ Snapshot API List

วิธีใช้:
    # จาก URL (server ต้องรันอยู่)
    python scripts/openapi_to_csv.py --url http://localhost:8000/openapi.json --output api_snapshot.csv

    # จากไฟล์
    python scripts/openapi_to_csv.py --file openapi.json --output api_snapshot.csv

    # Export เป็น Excel (.xlsx)
    python scripts/openapi_to_csv.py --url http://localhost:8000/openapi.json --output api_snapshot.xlsx

    # ดาวน์โหลด openapi.json ก่อน แล้วแปลง
    python scripts/openapi_to_csv.py --url http://localhost:8000/openapi.json --save-json openapi.json --output api_snapshot.csv
"""

import argparse
import csv
import json
import sys
from datetime import datetime
from pathlib import Path
from urllib.request import urlopen
from urllib.error import URLError


def fetch_openapi_json(url: str) -> dict:
    """ดึง openapi.json จาก URL"""
    try:
        with urlopen(url) as response:
            return json.loads(response.read().decode("utf-8"))
    except URLError as e:
        print(f"❌ ไม่สามารถเชื่อมต่อกับ {url}: {e}")
        sys.exit(1)


def load_openapi_file(filepath: str) -> dict:
    """โหลด openapi.json จากไฟล์"""
    path = Path(filepath)
    if not path.exists():
        print(f"❌ ไม่พบไฟล์: {filepath}")
        sys.exit(1)
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def extract_api_list(spec: dict) -> list[dict]:
    """แปลง OpenAPI spec ให้เป็น list ของ API endpoints"""
    rows = []
    paths = spec.get("paths", {})

    for path, methods in paths.items():
        for method, details in methods.items():
            if method in ("get", "post", "put", "patch", "delete", "head", "options"):
                # Extract parameters
                params = details.get("parameters", [])
                param_names = ", ".join(
                    f"{p.get('name', '?')} ({p.get('in', '?')})"
                    for p in params
                )

                # Extract request body schema
                request_body = ""
                rb = details.get("requestBody", {})
                if rb:
                    content = rb.get("content", {})
                    for content_type, schema_info in content.items():
                        ref = schema_info.get("schema", {}).get("$ref", "")
                        if ref:
                            request_body = ref.split("/")[-1]
                        else:
                            request_body = content_type

                # Extract response codes
                responses = details.get("responses", {})
                response_codes = ", ".join(sorted(responses.keys()))

                # Extract main response schema
                response_schema = ""
                success_response = responses.get("200") or responses.get("201") or {}
                if success_response:
                    content = success_response.get("content", {})
                    for content_type, schema_info in content.items():
                        ref = schema_info.get("schema", {}).get("$ref", "")
                        if ref:
                            response_schema = ref.split("/")[-1]

                # Extract tags
                tags = ", ".join(details.get("tags", []))

                rows.append({
                    "Method": method.upper(),
                    "Path": path,
                    "Summary": details.get("summary", ""),
                    "Description": (details.get("description", "") or "")[:200],
                    "Tags": tags,
                    "Parameters": param_names,
                    "Request Body": request_body,
                    "Response Codes": response_codes,
                    "Response Schema": response_schema,
                    "Operation ID": details.get("operationId", ""),
                    "Deprecated": "Yes" if details.get("deprecated") else "No",
                })

    return rows


def save_csv(rows: list[dict], output: str) -> None:
    """บันทึกเป็น CSV"""
    if not rows:
        print("⚠️  ไม่พบ API endpoints")
        return

    fieldnames = list(rows[0].keys())
    with open(output, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"✅ บันทึก CSV เรียบร้อย: {output} ({len(rows)} endpoints)")


def save_excel(rows: list[dict], output: str) -> None:
    """บันทึกเป็น Excel (.xlsx) — ต้องติดตั้ง openpyxl"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("❌ ต้องติดตั้ง openpyxl ก่อน: pip install openpyxl")
        print("   หรือใช้ --output api_snapshot.csv แทน")
        sys.exit(1)

    if not rows:
        print("⚠️  ไม่พบ API endpoints")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "API Endpoints"

    # ── Header Style ──
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ── Method colors ──
    method_colors = {
        "GET": PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid"),
        "POST": PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid"),
        "PUT": PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid"),
        "PATCH": PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid"),
        "DELETE": PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid"),
    }

    # Write header
    fieldnames = list(rows[0].keys())
    for col_idx, header in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, key in enumerate(fieldnames, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[key])
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            # Color-code Method column
            if key == "Method" and row_data[key] in method_colors:
                cell.fill = method_colors[row_data[key]]
                cell.font = Font(bold=True)

    # Auto-adjust column widths
    for col_idx, key in enumerate(fieldnames, 1):
        max_length = len(key)
        for row_data in rows:
            cell_value = str(row_data.get(key, ""))
            max_length = max(max_length, min(len(cell_value), 50))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_length + 4

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    wb.save(output)
    print(f"✅ บันทึก Excel เรียบร้อย: {output} ({len(rows)} endpoints)")


def main():
    parser = argparse.ArgumentParser(
        description="แปลง OpenAPI spec (openapi.json) จาก FastAPI ให้เป็น CSV หรือ Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
ตัวอย่าง:
  python scripts/openapi_to_csv.py --url http://localhost:8000/openapi.json --output api_snapshot.csv
  python scripts/openapi_to_csv.py --file openapi.json --output api_snapshot.xlsx
  python scripts/openapi_to_csv.py --url http://localhost:8000/openapi.json --save-json openapi.json --output api.csv
        """,
    )

    source = parser.add_mutually_exclusive_group(required=True)
    source.add_argument("--url", help="URL ของ openapi.json (เช่น http://localhost:8000/openapi.json)")
    source.add_argument("--file", help="Path ของไฟล์ openapi.json")

    parser.add_argument("--output", "-o", required=True, help="ไฟล์ output (.csv หรือ .xlsx)")
    parser.add_argument("--save-json", help="บันทึก openapi.json ไว้ด้วย (optional)")

    args = parser.parse_args()

    # ── Load spec ──
    print(f"📄 กำลังโหลด OpenAPI spec...")
    if args.url:
        spec = fetch_openapi_json(args.url)
        print(f"   จาก URL: {args.url}")
    else:
        spec = load_openapi_file(args.file)
        print(f"   จากไฟล์: {args.file}")

    # ── Print API info ──
    info = spec.get("info", {})
    print(f"\n📋 API: {info.get('title', 'Unknown')}")
    print(f"   Version: {info.get('version', 'Unknown')}")
    print(f"   Snapshot: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # ── Save JSON if requested ──
    if args.save_json:
        with open(args.save_json, "w", encoding="utf-8") as f:
            json.dump(spec, f, indent=2, ensure_ascii=False)
        print(f"\n💾 บันทึก openapi.json: {args.save_json}")

    # ── Extract & Save ──
    rows = extract_api_list(spec)
    print(f"\n🔍 พบ {len(rows)} API endpoints\n")

    output_path = args.output.lower()
    if output_path.endswith(".xlsx"):
        save_excel(rows, args.output)
    elif output_path.endswith(".csv"):
        save_csv(rows, args.output)
    else:
        print(f"❌ ไม่รองรับนามสกุลไฟล์: {args.output}")
        print("   รองรับ: .csv, .xlsx")
        sys.exit(1)


if __name__ == "__main__":
    main()
